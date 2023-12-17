import openpyxl
import sys
from mysqlConnection import getMyDB
import mysql.connector
import random


# Establish the database connection
mydb = getMyDB()
mydb.reconnect()
cursor = mydb.cursor()


def querySQL(query, values):
    try:
        cursor.execute(query, values)
    except mysql.connector.Error as e:
        print(e)
        return False
    return True


def insertLine(line, teamId):
    query = """UPDATE bowlTeams SET line = %s WHERE teamId = %s;"""
    values = (line, teamId)
    querySQL(query, values)


def generateId(prefix, length):
    output = str(prefix)
    for _ in range(length - len(output)):
        output += str(random.randint(0, 9))
    return int(output)


def insertPlayers(sheet):
    query = """INSERT INTO players (playerId, name, year) VALUES (%s, %s, %s)"""
    players = []
    for col in range(6, sheet.max_column + 1):
        name = sheet.cell(row=1, column=col).value
        id = generateId(2023, 9)
        players.append({"name": name, "id": id, "column": col})
        values = (id, name, 2023)
        # querySQL(query, values)
    return players


def insertPlayerPick(player, gameId, teamId):
    query = """INSERT INTO playerPicks (player, gameId, teamId, year) VALUES (%s, %s, %s, %s)"""
    values = (player, gameId, teamId, 2023)
    querySQL(query, values)


def insertPlayerPicksForRow(players, sheet, row, bowlId):
    teamId = sheet.cell(row=row, column=2).value
    for player in players:
        pick = sheet.cell(row=row, column=player["column"]).value
        if pick:
            insertPlayerPick(player["name"], bowlId, teamId)


def parse_excel(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        players = insertPlayers(sheet)
        row_count = sheet.max_row

        for row in range(2, row_count, 2):
            bowlId = sheet.cell(row=row + 1, column=1).value
            top = {
                "teamId": sheet.cell(row=row, column=2).value,
                "line": sheet.cell(row=row, column=5).value,
            }
            bottom = {
                "teamId": sheet.cell(row=row + 1, column=2).value,
                "line": sheet.cell(row=row + 1, column=5).value,
            }
            if top["line"]:
                insertLine(top["line"], top["teamId"])
            if bottom["line"]:
                insertLine(bottom["line"], bottom["teamId"])
            insertPlayerPicksForRow(players, sheet, row, bowlId)
            insertPlayerPicksForRow(players, sheet, row + 1, bowlId)

    except Exception as e:
        print(f"Error: {e}")
    finally:
        # Close the workbook
        workbook.close()


def main(args):
    excel_file_path = "resources/bowlpool-2023-picks.xlsx"
    parse_excel(excel_file_path)

    # Close the database connection
    cursor.close()
    mydb.commit()
    mydb.close()


if __name__ == "__main__":
    main(sys.argv)
